import os
import re
import glob as gl
import datetime as dt
import params as pr
import shutil
import whatstk as wa
import openpyxl
from openpyxl.styles import PatternFill


def get_txt_filenames(path, fullName=True, char=60):
    """
    Use this to get list to populate col A: filename
    :param path: string
    :return: Dict of all txt filenames in given path above. Key value refers to row number.
    """
    # Make sure we end with a trailing slash
    path = os.path.join(path, '')

    # If there's no path, return False
    if len(gl.glob(path+'*.txt')) == 0:
        return False
    else:
        # Get path/*someTxtFile*.txt, strip to just its filenames, make sure we don't get - Copy
        txt_paths = [os.path.basename(p) for p in gl.glob(path+'*.txt') if '- Copy' not in p]
        # Output dict, with key as row number for later.
        if fullName:
            return { list.index(txt_paths,p): p for p in txt_paths}
        else:
            return {list.index(txt_paths, p): p[:pr.FILENAME_CHAR_LENGTH] for p in txt_paths}


def get_txt_contents(path):
    """
    We use whatstk to create dataframe from raw chat logs. Makes it easy later for advance operation.
    :param path: string to where one wants to get chat log contents from
    :return: dict, key is row number, each value is dataframe of chat log
    """

    # Make sure we end with a trailing slash
    path = os.path.join(path, '')

    # Get filenames
    txts_names = get_txt_filenames(path)

    # Save contents into a dict. Each elemenet's key denotes row number, value is a pandas df containing content, with cols: date, username, message.
    # We check for only 2 cases: 24 hour clock and 12 hour clock with AM/PM. We assume no seconds.
    txts_contents = {}
    for row_num in txts_names:
        for date_format in pr.DATE_FORMATS:
            try:
                txts_contents[row_num] = wa.WhatsAppChat.from_source(path + txts_names[row_num], hformat=date_format).df
                break
            except:
                if pr.DATE_FORMATS.index(date_format)+1 == len(pr.DATE_FORMATS):
                    print('Unsupported date format in chat log:', txts_names[row_num])
                    break
                continue

    return txts_contents


def filter_date(contents):
    """
    Take in start and end date, keep chats that is within the range. Cut off all else.
    We assume that WhatsApp chat log always has oldest conversations at the top
    :param contents: dict, key is row number, each value is dataframe of chat log
    :return: contents: dict, key is row number, each value is dataframe of chat log
    """

    return {row_num: contents[row_num][(contents[row_num]['date'] > pr.START_DATE) &
            (contents[row_num]['date'] < pr.END_DATE)] for row_num in contents}


def standardize_datetime_formats(txt_filenames, source_path):

    print('Start of standardize_datetime_formats')

    REGEX_DATE_FORMATS = {0: "\w\w\w/\d\d*/\d\d,\s", 1: "\d\d*/\w\w\w/\d\d,\s", 2: "\w\w\w/\d\d*/\d\d\d\d,\s", 3: "\d\d*/\w\w\w/\d\d\d\d,\s",
                          4:"\d\d*/\d\d*/\d\d,\s", 5: "\d\d*\.\d\d*\.\d\d,\s", 6: "\d\d\d\d/\d\d*/\d\d*,\s",
                          7: "\w\w\w\.\d\d*\.\d\d\d\d,\s", 8: "\d\d*/\w\w\w/\d\d\d\d,\s"}

    ODD_DATE_FORMATS = {0: "%b/%d/%y, ", 1: "%d/%b/%y, ", 2: "%b/%d/%Y, ", 3: "%d/%b/%Y, ",
                        4:"%m/%d/%y, ", 5: "%m.%d.%y, ", 6: "%Y/%m/%d, ",
                        7: "%b.%d.%Y, ", 8: "%d.%b.%Y"}

    REGEX_TIME_FORMATS = {0: ", \d\d*:\d\d* \S\S - "}
    ODD_TIME_FORMAT = {0: ", %I:%M %p - "}

    for row_num in txt_filenames:
        # Open text file
        with open(os.path.join(source_path, txt_filenames[row_num]), 'rb') as f:
            txt_file = f.read().decode('utf-8')

        # Find each case of date format that we don't want, i.e. Mmm/DD/YY, MM/DD/YY, MM.DD.YY
        for fmt_idx in REGEX_DATE_FORMATS:
            # .findall returns a list ; if none found, list will have length of zero
            if len(re.findall(REGEX_DATE_FORMATS[fmt_idx], txt_file)) != 0:
                print('Found non standard date format:', REGEX_DATE_FORMATS[fmt_idx] ,'in chat log ', txt_filenames[row_num])
                txt_file_dates = re.findall(REGEX_DATE_FORMATS[fmt_idx], txt_file)

                # Replace with date format that we want, i.e. "%d/%m/%y, "  or "DD/MM/YYYY, "
                for date_in_txt in txt_file_dates:
                    do = dt.datetime.strptime(date_in_txt, ODD_DATE_FORMATS[fmt_idx])
                    txt_file = txt_file.replace(date_in_txt, do.strftime("%d/%m/%Y, "))

                with open(os.path.join(source_path,txt_filenames[row_num]) , "wb") as f:
                    f.write(txt_file.encode('utf-8'))
                break

        # Find each case of time format that we don't want, i.e. HH:MM AM , HH:MM PM
        for fmt_idx in REGEX_TIME_FORMATS:
            # .findall returns a list ; if none found, list will have length of zero
            if len(re.findall(REGEX_TIME_FORMATS[fmt_idx], txt_file)) != 0:
                print('Found non standard time format:', REGEX_TIME_FORMATS[fmt_idx], 'in chat log ', txt_filenames[row_num])
                txt_file_times = re.findall(REGEX_TIME_FORMATS[fmt_idx], txt_file)

                # Replace with time format that we want, i.e. HH:MM (24 hours)
                for time_in_txt in txt_file_times:
                    do = dt.datetime.strptime(time_in_txt, ODD_TIME_FORMAT[fmt_idx])
                    txt_file = txt_file.replace(time_in_txt, do.strftime(", %H:%M - "))

                with open(os.path.join(source_path,txt_filenames[row_num]) , "wb") as f:
                    f.write(txt_file.encode('utf-8'))
                break

    print('End of standardize_datetime_formats')
    return


def tidy_my_files(contents, txtfilenames, local_path_num, copyOnly):

    for row_num in txtfilenames:

        # Create folders
        tgt_folder_path = os.path.join(pr.TARGET_PATHS[local_path_num], '')
        if not os.path.exists(tgt_folder_path):
            os.mkdir(tgt_folder_path)
        tgt_chatlog_path = os.path.join(os.path.join(pr.TARGET_PATHS[local_path_num], '')+txtfilenames[row_num][:-4], '')
        if not os.path.exists(tgt_chatlog_path):
            os.mkdir(tgt_chatlog_path)

        # Move txt files
        source_chatlog_path = os.path.join(pr.SOURCE_PATHS[local_path_num], txtfilenames[row_num])
        if not copyOnly:
            shutil.move(source_chatlog_path, os.path.join(tgt_chatlog_path, txtfilenames[row_num]))
        if copyOnly:
            shutil.copy(source_chatlog_path, os.path.join(tgt_chatlog_path, txtfilenames[row_num]))

        # Move media files
        media_frRawMsg = {fmt: contents[row_num]['message'][contents[row_num]['message'].str.contains(fmt)].array
        if contents[row_num]['message'].str.contains(fmt).any()
        else None for fmt in pr.MEDIA_FORMATS}

        for fmt in media_frRawMsg:
            if media_frRawMsg[fmt] is not None:
                media_filenames = [msg[:msg.rfind(fmt)+len(fmt)] for msg in media_frRawMsg[fmt]]
                if not copyOnly:
                    move_it = [shutil.move(os.path.join(pr.SOURCE_PATHS[local_path_num], file), os.path.join(tgt_chatlog_path, file))
                                if os.path.exists(os.path.join(pr.SOURCE_PATHS[local_path_num], file))
                                else print('Found mention of file:', file ,'in chat log -', txtfilenames[row_num] , '; However, file not found in SOURCE_PATH')
                                for file in media_filenames]
                if copyOnly:
                    copy_it = [shutil.copy(os.path.join(pr.SOURCE_PATHS[local_path_num], file), os.path.join(tgt_chatlog_path, file))
                                if os.path.exists(os.path.join(pr.SOURCE_PATHS[local_path_num], file))
                                else print('Found mention of file:', file ,'in chat log -', txtfilenames[row_num] , '; However, file not found in SOURCE_PATH')
                                for file in media_filenames]

    return


def basic_format_excel(path):
    workbook = openpyxl.load_workbook(filename=path)
    worksheet = workbook.active
    # Adjust width of each col
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = length
    # Hide unncessary cols
    worksheet.column_dimensions.group(start=pr.HIDE_COL_START, end=pr.HIDE_COL_END, hidden=True)
    # Add background color to col names
    for rows in worksheet.iter_rows(min_row=1, max_row=1, min_col=2, max_col=10):
        for cell in rows:
            cell.fill = PatternFill(start_color=pr.COL_NAME_COLOUR, end_color=pr.COL_NAME_COLOUR, fill_type="solid")
    # Save excel
    workbook.save(path)
    return